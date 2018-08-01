"""
Created on 2018/8/1
@Author: Jeff Yang
"""
import openpyxl
import json

wb = openpyxl.Workbook()

ws = wb.active
ws['A1'] = '简历编号'
ws['B1'] = '简历最后更新时间'
ws['C1'] = '学历'
ws['D1'] = '性别'
ws['E1'] = '毕业院校'
ws['F1'] = '年龄'
ws['G1'] = '工作经历'

i = 1
with open('resume.json', 'r', encoding='utf-8') as f:
    while True:
        line = f.readline()
        data = []
        if line:
            dict = json.loads(line)
            data.append(dict['resume_id'])
            data.append(dict['last_update_time'])
            data.append(dict['education'])
            data.append(dict['gender'])
            data.append(dict['graduate_institution'])
            data.append(dict['age'])
            for item in dict['work_experiences']:
                data.append(item['company'])
                data.append(item['job_title'])
                data.append(item['date_range'])
                data.append(item['work_experience_describe'])
            ws.append(data)
            print('第', i, '条……')
            i += 1
        else:
            break

wb.save('resume.xlsx')
